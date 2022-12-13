import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import doctest

class DataVac:
    """
    класс Вакансии
    """
    def __init__(self, name, minSalary, maxSalary, currSalary, area_name, published_at):
        """
        :param name (str): название вакансии
        :param minSalary (int): нижняя граница вилки оклада
        :param maxSalary (int): верхняя граница вилки оклада
        :param currSalary (str): валюта ЗП
        :param area_name (str): локация
        :param published_at (str): время публикации
        """
        self.name = name
        self.salary = Salary(minSalary, maxSalary, currSalary)
        self.area_name = area_name
        self.published_at = published_at

class Salary:
    """
    класс для представления зарплаты
    """
    def __init__(self, minSalary, maxSalary, currSalary):
        """

        :param minSalary (int): нижняя граница вилки оклада
        :param maxSalary (int): верхняя граница вилки оклада
        :param currSalary (str): валюта ЗП
        """
        self.minSalary = int(float(minSalary))
        self.maxSalary = int(float(maxSalary))
        self.currSalary = currSalary

    def rubleConverter(self):
        """
        метод, для конвертации средней зп в рубли
        :return: средняя ЗП умноженная на курс рубля
        """
        return (self.minSalary + self.maxSalary) / 2 * self.__rubleCourse[self.currSalary]

    __rubleCourse = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }


class InputConnection:
    """
    класс, начинающий работу программы тем, что запрашивает input пользователя
    """
    def dataInput(self):
        """
        метод, по сути, выполняющий всю работу ( вызов всех необходимых методов )
        :return: статистика
        """
        fileName = input('Введите название файла: ')

        jobName = input('Введите название профессии: ')

        rubleSal = self.__createDict()
        salCount = self.__createDict()

        rubleJob = self.__createDict()
        jobCounter = self.__createDict()

        dVacs = []

        with open(fileName, encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = []

            isFirstEl = True
            for r in reader:
                if isFirstEl:
                    isFirstEl = False
                    header = r
                else:
                    if not "" in r and len(r) == len(header):

                        dv = DataVac(
                            r[header.index('name')],
                            r[header.index('salary_from')],
                            r[header.index('salary_to')],
                            r[header.index('salary_currency')],
                            r[header.index('area_name')],
                            r[header.index('published_at')]
                        )
                        dVacs.append(dv)

                        yearDate = int(dv.published_at[:4])
                        rubleSal[yearDate] = self.__avg(rubleSal[yearDate], dv.salary.rubleConverter(),
                                                         salCount[yearDate])
                        salCount[yearDate] += 1

                        if (dv.name.find(jobName) != -1):
                            rubleJob[yearDate] = self.__avg(rubleJob[yearDate], dv.salary.rubleConverter(), jobCounter[yearDate])
                            jobCounter[yearDate] += 1

        rubleSal = self.__erEmpty(self.__roundVals(rubleSal))
        rubleJob = self.__erEmpty(self.__roundVals(rubleJob))
        salCount = self.__erEmpty(salCount)
        jobCounter = self.__erEmpty(jobCounter)

        print('Динамика уровня зарплат по годам:', rubleSal)
        print('Динамика количества вакансий по годам:', salCount)
        print('Динамика уровня зарплат по годам для выбранной профессии:', rubleJob)
        print('Динамика количества вакансий по годам для выбранной профессии:', jobCounter)

        citiesSal = {}
        citiesCount = {}
        citiesFrac = {}

        for vac in dVacs:
            city = vac.area_name
            if city not in citiesSal.keys():
                if len([x for x in dVacs if x.area_name == city]) >= int(len(dVacs) / 100):
                    citiesSal[city] = vac.salary.rubleConverter()
                    citiesCount[city] = 1
            else:
                citiesSal[city] = self.__avg(citiesSal[city], vac.salary.rubleConverter(), citiesCount[city])
                citiesCount[city] += 1

        summary = len(dVacs)
        for k, v in citiesCount.items():
            citiesFrac[k] = round(v / (summary / 100) / 100, 4)

        citiesSal = self.__roundVals(self.__erEmpty(self.__sortCities(citiesSal)))
        citiesFrac = self.__erEmpty(self.__sortCities(citiesFrac))

        print('Уровень зарплат по городам (в порядке убывания):', citiesSal)
        print('Доля вакансий по городам (в порядке убывания):', citiesFrac)

        return jobName, rubleSal, salCount, rubleJob, jobCounter, citiesSal, citiesFrac

    def __sortCities(self, d):
        """
        сортирует словарь
        :param d (dict): словарь городов
        :return: отсортированный словарь
        """
        return dict(sorted(d.items(), key=lambda x: x[1], reverse=True)[:10])

    def __roundVals(self, d):
        """
        :param d: словарь значений
        :return (dict): словарь округленных значений
        """
        return dict(map(lambda x: (x[0], int(x[1])), d.items()))

    def __createDict(self):
        """
        :return (dict): словарь с ключами от 2007 до 2023
        """
        return {x: 0 for x in range(2007, 2023)}

    def __erEmpty(self, d):
        cd = dict(filter(lambda x: x[1], d.items()))
        if len(cd.keys()) == 0:
            cd[2022] = 0
        return cd

    def __avg(self, m, x, n):
        """
        :return (int): возвращает среднее значение
        """
        return (m * n + x) / (n + 1)



class Report:
    """
    класс, ответственный за формирование отчета
    """
    def __init__(self):
        self.wBook = Workbook()

        for elementName in self.wBook.sheetnames:
            element = self.wBook[elementName]
            self.wBook.remove(element)

        self.wBook.create_sheet('Статистика по годам')
        self.wBook.create_sheet('Статистика по городам')

    __startHeaders = [
        'Год',
        'Средняя зарплата',
        'Средняя зарплата - ',
        'Количество вакансий',
        'Количество вакансий - '
    ]

    def __createFirstSheet(self, data):
        """
        создает первую строку отчета статистика по годам
        :param data: статистика по годам
        """
        self.wBook.active = self.wBook['Статистика по годам']
        wActive = self.wBook.active

        self.__startHeaders[2] = self.__startHeaders[2] + data[0]
        self.__startHeaders[4] = self.__startHeaders[4] + data[0]
        wActive.append(vac for vac in self.__startHeaders)
        for r in wActive.rows:
            for c in r:
                c.font = Font(bold=True)

        for yearDate in data[1].keys():
            r = [yearDate, data[1][yearDate], data[3][yearDate], data[2][yearDate], data[4][yearDate]]
            wActive.append(r)

        self.__sizeSetter()
        self.__createBorder()

    def __sizeSetter(self):
        """
        устанавливает размер ячейки
        """
        for colCells in self.wBook.active.columns:
            length = max(len(self.__getAsText(cell.value)) for cell in colCells)
            self.wBook.active.column_dimensions[get_column_letter(colCells[0].column)].width = length + 2

    def __getAsText(self, val):
        """
        возвращает как строку
        :param val: значение
        :return: строковое значение
        """
        if val is None:
            return ""
        return str(val)

    def __createBorder(self):
        """
        создает границу
        """
        for r in self.wBook.active.rows:
            for c in r:
                c.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    __lastHeaders = [
        'Город',
        'Уровень зарплат',
        '',
        'Город',
        'Доля вакансий'
    ]

    def generate_excel(self, dataFirst, dataSecond):
        """
        создает эксель документ
        :param dataFirst: статистика по годам
        :param dataSecond: статистика по городам
        """
        self.__createFirstSheet(dataFirst)
        self.__createSecondSheet(dataSecond)

        self.wBook.save('report.xlsx')

    def __createSecondSheet(self, data):
        """
        создает вторую строку отчета статистика по городам
        :param data: статистика по городам
        """
        self.wBook.active = self.wBook['Статистика по городам']
        wActive = self.wBook.active

        wActive.append(v for v in self.__lastHeaders)
        for r in wActive.rows:
            for c in r:
                c.font = Font(bold=True)

        data1 = list(data[0].keys())
        data2 = list(data[0].values())
        data3 = list(data[1].keys())
        data4 = list(data[1].values())

        for i in range(len(data[0])):
            row = [data1[i], data2[i], '', data3[i], data4[i]]
            wActive.append(row)

        self.__sizeSetter()
        self.__createBorder()

        for i in range(1, 12):
            wActive[f"C{i}"].border = Border()

        for i in range(1, 12):
            wActive[f"E{i}"].number_format = '0.00%'

        self.wBook.active = self.wBook['Статистика по годам']


inpConnection = InputConnection()
report = Report()
data = inpConnection.dataInput()
report.generate_excel(data[:5], data[5:])
